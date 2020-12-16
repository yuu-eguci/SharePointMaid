import datetime
import calendar
import openpyxl
import util

# このモジュール用のロガーを取得します。
logger = util.get_my_logger(__name__)


def run(workbook_path: str) -> None:
    """指定した workbook に M月第N月曜日 シートを追加します。

    Args:
        workbook_path (str): workbook のパス
    """

    # 対象 workbook 取得します。
    workbook = openpyxl.load_workbook(workbook_path)

    # 来週の月曜日を取得し、 M月第N月曜日 を取得します。
    next_monday = get_next_monday()
    dow = get_nth_dow(next_monday.year, next_monday.month, next_monday.day)
    sheet_title = f'{next_monday.month}月第{dow[0]}月曜日'

    # もしそのシート追加済みならやることありません。
    logger.debug(workbook.sheetnames)
    if sheet_title in workbook.sheetnames:
        logger.info(f'{sheet_title} はもう存在します。コピーはスキップします。')
        return

    # Format base という sheet を複製して、 MM月第N月曜という名前で最後に追加します。
    # NOTE: Sheet が存在しないと KeyError が発生。
    copied_sheet = workbook.copy_worksheet(workbook['Format base'])
    copied_sheet.title = sheet_title
    workbook.save(workbook_path)


def get_next_monday() -> datetime.date:
    """次の月曜を datetime.date オブジェクトで取得します。

    Returns:
        datetime.date: 次の月曜(Ex: datetime.date(2020, 12, 21))
    """

    # Ex: datetime.date(2020, 12, 16)
    today = datetime.date.today()

    # 今日からの次の月曜への日数を取得します。
    # NOTE: 次の月曜(0)を取得するロジックです。
    shift_num = 0 - today.weekday()
    shift_num = shift_num + 7 if shift_num < 0 else shift_num
    delta = datetime.timedelta(days=shift_num)
    next_monday = today + delta
    return next_monday


def get_nth_week(day: int) -> int:
    """第何週、取得します。

    Args:
        day (int): day

    Returns:
        int: 第何週
    """

    return (day - 1) // 7 + 1


def get_nth_dow(year: int, month: int, day: int) -> tuple:
    """(第何週, 何曜日)を取得します。

    Args:
        year (int): datetime.date.year
        month (int): datetime.date.month
        day (int): datetime.date.day

    Returns:
        tuple: (第何週, 何曜日)
    """

    return get_nth_week(day), calendar.weekday(year, month, day)


if __name__ == '__main__':
    logger.info('Next Monday')
    next_monday = get_next_monday()
    logger.info(repr(next_monday))
    dow = get_nth_dow(next_monday.year, next_monday.month, next_monday.day)
    logger.info(dow)
    logger.info(f'{repr(next_monday)} は {next_monday.month}月第{dow[0]}月曜日')
    run('./sample.xlsx')

